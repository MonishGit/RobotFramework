import openpyxl

wk = openpyxl.load_workbook('C:\Users\mopawar\PycharmProjects\Creating_NewFramework\Myexcel.xlsx')
sh = wk['Test3']

__all__ = ['find_max_rows','find_cell_data']
# Create function to identify max rows and Cell data in spreadsheet ----THIS IS VERY IMP

# Step 1 Max rows

# class ExcelExtract ():

def find_max_rows(self,sheetname1):
    sh1 = wk[sheetname1]
    return sh1.max_row

    # Step 2 identify cell data

# def find_cell_data(self,sheetname1, row, column)
def find_cell_data (self, sheetname,row,column):
    sh1 = wk[sheetname]
    cell_data = unicode(sh1.cell(row,column).value)
    return  cell_data

obj= ExcelExtract()

print   obj.find_cell_data('Test3',1,1)
print   obj.find_max_rows('Test3')