#import excel library

# Issue faced while loading libray- I installed openpyxl but not set right interpreter so it was showing me error -No module named openpyxl.
# I went in Edit run and installd openpy for this project interpreter (vertual env). File > settings > project intepreter >  +

import  openpyxl

#Featching workbook

myxl = openpyxl.load_workbook("C:\Users\mopawar\PycharmProjects\Creating_NewFramework\Myexcel.xlsx")
print (myxl.sheetnames)

#Navigating to specific sheet, in below example navigating to 'Test1'

print(myxl.active.title)
sheetname= myxl['Test3']

#Read the data from perticular cells

# print(sheetname['A3'].value)

#OR

c1 = sheetname.cell(5,3)
print(c1.value)

#show all data from excel by identifying maxrow and column

colnum= sheetname.max_column
rownum= sheetname.max_row

print(colnum)
print(rownum)

# for i in range(1,rownum+1):
#     for c in range(1,colnum+1):
#         d=sheetname.cell(i,c)
#         print (d.value)

# show all data from excel using complete table range

# for i in sheetname['A1':'C5']:
#     for c in i :
#         print (c.value)



